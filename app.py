import os
import json
import logging
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import tempfile

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Create Flask app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev_secret_key_for_replit")
CORS(app)

def get_gemini_model(api_key):
    genai.configure(api_key=api_key)
    return genai.GenerativeModel("models/gemini-2.5-flash")




@app.route('/')
def dashboard():
    """Home dashboard displaying predefined topics"""
    return render_template('dashboard.html')

@app.route('/generator/<topic_id>')
def generator(topic_id):
    """Content generation interface for specific topic"""
    return render_template('generator.html', topic_id=topic_id)

@app.route('/api/generate-content', methods=['POST'])
def generate_content():
    """Generate educational content using Gemini Pro"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        api_key = data.get('api_key')
        prompt = data.get('prompt')
        temperature = float(data.get('temperature', 0.7))
        content_type = data.get('content_type', 'mcq')

        if not api_key:
            return jsonify({'error': 'Gemini API key is required'}), 400
        if not prompt:
            return jsonify({'error': 'Prompt is required'}), 400

        # ✅ Initialize Gemini model
        model = get_gemini_model(api_key)
        enhanced_prompt = enhance_prompt_for_content_type(prompt, content_type)

        # ✅ Generate response
        response = model.generate_content(enhanced_prompt)

# Gemini 2.5 responses
        if response.candidates and response.candidates[0].content.parts:
            response_text = response.candidates[0].content.parts[0].text
        else:
            raise ValueError("Empty or invalid response from Gemini")

        content = json.loads(response_text)


        return jsonify({
            'success': True,
            'content': content,
            'content_type': content_type
        })

    except Exception as e:
        logging.error(f"Error generating content: {e}")
        return jsonify({'error': f'Failed to generate content: {str(e)}'}), 500


# ✅ Enhance Prompt Function
def enhance_prompt_for_content_type(prompt, content_type):
    """Enhance the prompt based on content type for better structured output"""
    content_type_instructions = {
        'mcq': '''
Generate multiple choice questions with the following JSON structure:
{
    "title": "Topic Title",
    "questions": [
        {
            "question": "Question text",
            "options": ["A) Option 1", "B) Option 2", "C) Option 3", "D) Option 4"],
            "correct_answer": "A",
            "explanation": "Explanation for the correct answer"
        }
    ]
}
''',
        'cheat_sheet': '''
Generate a comprehensive cheat sheet with the following JSON structure:
{
    "title": "Cheat Sheet Title",
    "sections": [
        {
            "heading": "Section Title",
            "content": [
                "Key point 1",
                "Key point 2",
                "Key point 3"
            ]
        }
    ],
    "quick_tips": ["Tip 1", "Tip 2", "Tip 3"]
}
''',
        'drag_drop': '''
Generate drag and drop matching activities with the following JSON structure:
{
    "title": "Activity Title",
    "instructions": "Match the items on the left with the correct items on the right",
    "items": [
        {
            "left": "Item to match",
            "right": "Correct match"
        }
    ]
}
''',
        'textual': '''
Generate textual questions and exercises with the following JSON structure:
{
    "title": "Exercise Title",
    "questions": [
        {
            "type": "short_answer",
            "question": "Question text",
            "sample_answer": "Expected answer or key points"
        },
        {
            "type": "essay",
            "question": "Essay question",
            "guidelines": "Guidelines for answering"
        }
    ]
}
''',
        'listening': '''
Generate listening exercise content with the following JSON structure:
{
    "title": "Listening Exercise Title",
    "audio_description": "Description of the audio content (since actual audio cannot be generated)",
    "pre_listening": [
        "Pre-listening question 1",
        "Pre-listening question 2"
    ],
    "while_listening": [
        {
            "question": "While listening question",
            "type": "multiple_choice",
            "options": ["A) Option 1", "B) Option 2", "C) Option 3"],
            "correct_answer": "A"
        }
    ],
    "post_listening": [
        "Post-listening discussion question 1",
        "Post-listening discussion question 2"
    ]
}
'''
    }

    instruction = content_type_instructions.get(content_type, content_type_instructions['mcq'])
    return f"{prompt}\n\n{instruction}\n\nEnsure the response is valid JSON format."


# ✅ Export to Excel
@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export generated content to Excel file"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        content = data.get('content')
        content_type = data.get('content_type')
        topic_title = data.get('topic_title', 'English Course Content')

        if not content:
            return jsonify({'error': 'No content to export'}), 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Course Content"

        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"{topic_title} - {content_type.upper()}"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        if content_type == 'mcq':
            export_mcq_to_excel(ws, content, row, header_font, normal_font, header_fill)
        elif content_type == 'cheat_sheet':
            export_cheat_sheet_to_excel(ws, content, row, header_font, normal_font, header_fill)
        elif content_type == 'drag_drop':
            export_drag_drop_to_excel(ws, content, row, header_font, normal_font, header_fill)
        elif content_type == 'textual':
            export_textual_to_excel(ws, content, row, header_font, normal_font, header_fill)
        elif content_type == 'listening':
            export_listening_to_excel(ws, content, row, header_font, normal_font, header_fill)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(excel_file.getvalue())
        temp_file.close()

        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"{topic_title.replace(' ', '_')}_{content_type}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        return jsonify({'error': f'Failed to export to Excel: {str(e)}'}), 500


# ✅ Export to JSON
@app.route('/api/export-json', methods=['POST'])
def export_json():
    """Export generated content to JSON file"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        content = data.get('content')
        content_type = data.get('content_type')
        topic_title = data.get('topic_title', 'English Course Content')

        if not content:
            return jsonify({'error': 'No content to export'}), 400

        json_export = {
            'metadata': {
                'title': topic_title,
                'content_type': content_type,
                'generated_at': datetime.now().isoformat(),
                'application': 'English Communication Course Builder'
            },
            'content': content
        }

        json_content = json.dumps(json_export, indent=2, ensure_ascii=False)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.json', mode='w', encoding='utf-8')
        temp_file.write(json_content)
        temp_file.close()

        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"{topic_title.replace(' ', '_')}_{content_type}.json",
            mimetype='application/json'
        )

    except Exception as e:
        logging.error(f"Error exporting to JSON: {e}")
        return jsonify({'error': f'Failed to export to JSON: {str(e)}'}), 500


# Excel helper functions (unchanged)
def export_mcq_to_excel(ws, content, start_row, header_font, normal_font, header_fill):
    row = start_row
    headers = ['Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer', 'Explanation']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for q in content.get('questions', []):
        ws.cell(row=row, column=1, value=q.get('question', '')).font = normal_font
        options = q.get('options', [])
        for i, option in enumerate(options[:4], 2):
            ws.cell(row=row, column=i, value=option).font = normal_font
        ws.cell(row=row, column=6, value=q.get('correct_answer', '')).font = normal_font
        ws.cell(row=row, column=7, value=q.get('explanation', '')).font = normal_font
        row += 1

def export_cheat_sheet_to_excel(ws, content, start_row, header_font, normal_font, header_fill):
    row = start_row
    sections = content.get('sections', [])
    for section in sections:
        ws.merge_cells(f'A{row}:C{row}')
        heading_cell = ws[f'A{row}']
        heading_cell.value = section.get('heading', '')
        heading_cell.font = header_font
        heading_cell.fill = header_fill
        row += 1
        for item in section.get('content', []):
            ws.cell(row=row, column=1, value=f"• {item}").font = normal_font
            row += 1
        row += 1
    if content.get('quick_tips'):
        ws.merge_cells(f'A{row}:C{row}')
        tips_cell = ws[f'A{row}']
        tips_cell.value = "Quick Tips"
        tips_cell.font = header_font
        tips_cell.fill = header_fill
        row += 1
        for tip in content.get('quick_tips', []):
            ws.cell(row=row, column=1, value=f"• {tip}").font = normal_font
            row += 1

def export_drag_drop_to_excel(ws, content, start_row, header_font, normal_font, header_fill):
    row = start_row
    ws.merge_cells(f'A{row}:B{row}')
    inst_cell = ws[f'A{row}']
    inst_cell.value = content.get('instructions', '')
    inst_cell.font = header_font
    row += 2
    ws.cell(row=row, column=1, value='Left Item').font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value='Correct Match').font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    row += 1
    for item in content.get('items', []):
        ws.cell(row=row, column=1, value=item.get('left', '')).font = normal_font
        ws.cell(row=row, column=2, value=item.get('right', '')).font = normal_font
        row += 1

def export_textual_to_excel(ws, content, start_row, header_font, normal_font, header_fill):
    row = start_row
    headers = ['Type', 'Question', 'Sample Answer/Guidelines']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for q in content.get('questions', []):
        ws.cell(row=row, column=1, value=q.get('type', '')).font = normal_font
        ws.cell(row=row, column=2, value=q.get('question', '')).font = normal_font
        answer_or_guidelines = q.get('sample_answer', '') or q.get('guidelines', '')
        ws.cell(row=row, column=3, value=answer_or_guidelines).font = normal_font
        row += 1

def export_listening_to_excel(ws, content, start_row, header_font, normal_font, header_fill):
    row = start_row
    ws.merge_cells(f'A{row}:C{row}')
    audio_cell = ws[f'A{row}']
    audio_cell.value = f"Audio Description: {content.get('audio_description', '')}"
    audio_cell.font = header_font
    row += 2
    if content.get('pre_listening'):
        ws.cell(row=row, column=1, value='Pre-listening Questions').font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        row += 1
        for q in content.get('pre_listening', []):
            ws.cell(row=row, column=1, value=f"• {q}").font = normal_font
            row += 1
        row += 1
    if content.get('while_listening'):
        ws.cell(row=row, column=1, value='While Listening Questions').font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        row += 1
        for q in content.get('while_listening', []):
            ws.cell(row=row, column=1, value=q.get('question', '')).font = normal_font
            if q.get('options'):
                ws.cell(row=row, column=2, value=' | '.join(q.get('options', []))).font = normal_font
            if q.get('correct_answer'):
                ws.cell(row=row, column=3, value=f"Answer: {q.get('correct_answer')}").font = normal_font
            row += 1
        row += 1
    if content.get('post_listening'):
        ws.cell(row=row, column=1, value='Post-listening Questions').font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        row += 1
        for q in content.get('post_listening', []):
            ws.cell(row=row, column=1, value=f"• {q}").font = normal_font
            row += 1




@app.route('/api/models')
def list_models():
    api_key = request.args.get('api_key')
    if not api_key:
        return jsonify({'error': 'Missing api_key query param'}), 400

    genai.configure(api_key=api_key)
    models = genai.list_models()
    return jsonify({'available_models': [m.name for m in models]})



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
